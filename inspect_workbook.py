import openpyxl
from pathlib import Path

path = Path(r'd:\tns-hrms\MoWCA_Officers_List_by_Department (1).xlsx')
wb = openpyxl.load_workbook(path, data_only=True)
print('SHEETS', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print('\n===', name, '===')
    print('rows', ws.max_row, 'cols', ws.max_column)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
        print(row)
